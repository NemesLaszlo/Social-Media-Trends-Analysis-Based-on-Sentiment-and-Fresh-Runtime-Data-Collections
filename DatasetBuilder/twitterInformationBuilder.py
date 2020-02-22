import openpyxl
import operator
import os
import time
import tweepy
import urllib.request


class twitterInformationBuilder:

    def __init__(self, ConsumerAPI_Key, ConsumerAPI_Secret, Access_Token, Access_Token_Secret):
        self.consumerAPI_Key = ConsumerAPI_Key
        self.consumerAPI_Secret = ConsumerAPI_Secret
        self.accessToken = Access_Token
        self.accessTokenSecret = Access_Token_Secret

        auth = tweepy.OAuthHandler(self.consumerAPI_Key, self.consumerAPI_Secret)
        auth.set_access_token(self.accessToken, self.accessTokenSecret)
        self.api = tweepy.API(auth)

    def create_directory(self, dir_name):

        if not os.path.exists("data"):
            try:
                os.mkdir("data")
                print("Created directory 'data'")
            except:
                print("Unable to create directory 'data': Directory already exists")
        else:
            print("Unable to create directory 'data': Directory already exists")

        if not os.path.exists("data/data_" + dir_name):
            try:
                os.mkdir("data/data_" + dir_name)
                print("Created directory 'data/data_" + dir_name + "'")
            except:
                print("Unable to create directory 'data/data_" + dir_name + "': Directory already exists")
        else:
            print("Unable to create directory 'data/data_" + dir_name + "': Directory already exists")

        if not os.path.exists("data/data_" + dir_name + '/img'):
            try:
                os.mkdir("data/data_" + dir_name + '/img')
                print("Created directory 'data/data_" + dir_name + "/img'")
            except:
                print("Unable to create directory 'data/data_" + dir_name + "/img': Directory already exists")
        else:
            print("Unable to create directory 'data/data_" + dir_name + "/img': Directory already exists")

    def information_builder(self, tag, limit, lang):

        self.create_directory(dir_name=tag)

        print("Information Build Start")
        file_path = "data/data_" + tag

        tweets = tweepy.Cursor(self.api.search, q=tag, lang=lang).items(limit)

        # Excel setup
        tag_File = file_path + "/" + tag + "_Twitter.xlsx"
        wb = openpyxl.Workbook()
        ws_Captions = wb.create_sheet(title="Tweets")
        col = 'A'
        row = 1

        img_src = []
        hashtags = {}
        ext_links = []

        # Tweet workflow, tweets, hashtags and links separation,sorting and methodize
        for tweet in tweets:
            text = tweet.text.lower()
            ws_Captions[col + str(row)] = text

            for tag in text.split():
                if tag.startswith("#"):
                    if tag[1:] not in hashtags:
                        hashtags[tag[1:]] = 1
                    elif tag[1:] in hashtags:
                        hashtags[tag[1:]] = hashtags[tag[1:]] + 1

                if tag[:4] == 'http':
                    ext_links.append(tag)

            try:
                media = tweet.entities['media'][0]['media_url']
                if len(media) > 1:
                    img_src.append(media)
            except:
                pass

            row += 1

        hashtags = sorted(hashtags.items(), key=operator.itemgetter(1), reverse=True)

        ws_Tags = wb.create_sheet(title="Tags")
        tagName = 'A'
        tagFreq = 'B'
        row = 1

        for tag in hashtags:
            ws_Tags[tagName + str(row)] = tag[0]
            ws_Tags[tagFreq + str(row)] = tag[1]
            row += 1

        ws_Links = wb.create_sheet(title="Links")
        row = 1
        for link in ext_links:
            ws_Links['A' + str(row)] = link
            row += 1

        wb.save(tag_File)
        time.sleep(5)
        self.image_management(file_path=file_path, img_src=img_src)

    def image_management(self, file_path, img_src):
        row = 1
        for src in img_src:
            try:
                urllib.request.urlretrieve(src, file_path + '/img/Twitter_' + str(row) + ".jpeg")
                row += 1
                time.sleep(1.5)
            except:
                print("Image Download Failed. Downloading next image")
